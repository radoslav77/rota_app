"""
Parser for the hotel kitchen rota Excel format.
Extracts staff, sections, and shift entries from uploaded .xlsx files.
"""
import openpyxl
from datetime import datetime, date
import re

SECTION_HEADERS = {
    'EXECUTIVE CHEF', 'CONFERENCE & EVENTS', 'DUTY CHEFS',
    'MAIN KITCHEN SAUCE SECTION', 'MAIN KITCHEN GARNISH SECTION',
    'MAIN KITCHEN LARDER', 'BREAKFAST SHIFT', 'NIGHT SHIFT',
    'PARK LANE CAFÉ "STAFF CANTEEN"', 'PARK LANE CAFE "STAFF CANTEEN"',
    'PASTRY & BAKERY', 'CHEF SUPPORT', 'SUPPORT CHEFS',
}

NON_SHIFT_VALUES = {
    'OFF', 'H', 'SICK', 'Comp', 'Paternity', 'Maternity',
    'TBC', 'OFF/R', 'No Work', '', None,
}

DAY_COLS = list(range(1, 15))  # columns B-O (0-indexed 1..14)


def is_section_header(val):
    if not val or not isinstance(val, str):
        return False
    return val.upper() in {h.upper() for h in SECTION_HEADERS} or val.upper() in SECTION_HEADERS


def is_staff_row(row):
    """A staff row has a name in col A and shift data in cols B-O"""
    val = row[0]
    if not val or not isinstance(val, str):
        return False
    if is_section_header(val):
        return False
    if val.upper() in ('CHEF NAME', 'UP COMING HIGHLIGHTS', 'UPCOMING HIGHLIGHTS'):
        return False
    if val.startswith('`') or val.startswith(']'):
        return False
    # Check if any day col has a shift-like value
    for i in range(1, 15):
        if i < len(row) and row[i] and isinstance(row[i], str):
            if '-' in row[i] or row[i].upper() in ('OFF', 'H', 'SICK', 'TBC', 'COMP'):
                return True
    return False


def extract_name_role(cell_val):
    """Split 'John Smith CHEF DE PARTIE' into name and role"""
    if not cell_val:
        return '', ''
    val = str(cell_val).strip()
    # Known role suffixes
    roles = [
        'EXECUTIVE CHEF', 'SOUS CHEF', 'JR. SOUS CHEF', 'JR SOUS CHEF',
        'CHEF DE PARTIE', 'DEMI CHEF DE PARTIE', 'DEMI CHEF',
        'COMMIS CHEF', 'NIGHT CHEF DE PARTIE', 'PASTRY CHEF',
        'NIGHT BAKER', 'SOUS', 'PART TIME', '(PART TIME)',
    ]
    role = ''
    name = val
    for r in sorted(roles, key=len, reverse=True):
        idx = val.upper().find(r)
        if idx > 0:
            name = val[:idx].strip()
            role = val[idx:].strip()
            break
    return name, role


def parse_rota_sheet(ws):
    """
    Returns:
        dates: list of date objects (14 dates, cols B-O)
        sections_data: list of {section, staff_rows}
        highlights: list of strings
    """
    all_rows = list(ws.iter_rows(values_only=True))
    
    # Find date row (row 3 in the file = index 2)
    dates = []
    date_row_idx = None
    for idx, row in enumerate(all_rows):
        if row[1] and isinstance(row[1], datetime):
            dates = []
            for i in range(1, 15):
                val = row[i] if i < len(row) else None
                if isinstance(val, datetime):
                    dates.append(val.date())
                elif isinstance(val, date):
                    dates.append(val)
                else:
                    dates.append(None)
            date_row_idx = idx
            break

    # Collect highlights
    highlights = []
    for row in all_rows[:7]:
        val = row[0]
        if val and isinstance(val, str) and 'HIGHLIGHTS' in val.upper():
            for i in range(1, 15):
                if i < len(row) and row[i]:
                    highlights.append(str(row[i]))

    # Parse sections and staff
    sections_data = []
    current_section = 'GENERAL'
    
    for row_idx, row in enumerate(all_rows):
        val = row[0]
        if not val:
            continue
        
        if isinstance(val, str) and is_section_header(val):
            current_section = val.strip()
            continue
        
        # Check if formula referencing another cell (section ref)
        if isinstance(val, str) and val.startswith('='):
            continue

        if is_staff_row(row):
            name, role = extract_name_role(val)
            if not name:
                continue
            shifts = {}
            for col_i, d in enumerate(dates):
                if d is None:
                    continue
                raw = row[col_i + 1] if (col_i + 1) < len(row) else None
                if raw is None:
                    shift_val = 'OFF'
                elif isinstance(raw, str):
                    shift_val = raw.strip() or 'OFF'
                else:
                    shift_val = str(raw).strip() or 'OFF'
                shifts[d.isoformat()] = shift_val

            sections_data.append({
                'section': current_section,
                'name': name,
                'role': role,
                'shifts': shifts,
            })

    return dates, sections_data, highlights


def parse_workbook(file_path):
    """Parse all rota sheets from workbook. Returns list of period dicts."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    rota_sheets = [s for s in wb.sheetnames if s not in ('overtime', 'cross charge salary Jan', '1', '2')]
    
    results = []
    for sheet_name in rota_sheets:
        ws = wb[sheet_name]
        dates, sections_data, highlights = parse_rota_sheet(ws)
        if dates and sections_data:
            results.append({
                'label': sheet_name,
                'dates': dates,
                'sections_data': sections_data,
                'highlights': ' | '.join(highlights),
            })
    return results
