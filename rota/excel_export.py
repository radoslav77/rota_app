"""
Export a generated rota back to the same .xlsx format as the original file.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# ── Colour palette (ARGB format for openpyxl) ────────────────────────────────
# Section header colours matching the ODS reference file
SECTION_HEADER_COLORS = {
    'EXECUTIVE':   'FFFFD966',   # amber/gold
    'CONFERENCE':  'FF00B050',   # bright green
    'DUTY':        'FFAEAAAA',   # mid grey
    'SAUCE':       'FFFF0000',   # red
    'GARNISH':     'FFB4C6E7',   # steel blue
    'LARDER':      'FF92D050',   # lime green
    'BREAKFAST':   'FFFFE699',   # pale yellow
    'NIGHT':       'FF333333',   # near-black
    'CANTE':       'FF808080',   # grey
    'CAFE':        'FF808080',
    'PASTRY':      'FF00B0F0',   # sky blue
    'TERRACE':     'FFF4B084',   # salmon/peach
}

# Shift cell colours (time-band based, matching ODS)
COLOR_EARLY     = 'FFC0E6F5'   # before 08:00 — pale sky
COLOR_AM        = 'FFE2EFDA'   # 08:00–11:59  — light green
COLOR_BREAKFAST = 'FFC4F2DE'   # breakfast shifts (#c4f2de)
COLOR_BRUNCH    = 'FF457AED'   # brunch shifts  (#457aed)
COLOR_PM      = 'FFFCE4D6'   # 12:00–17:59  — peach
COLOR_LATE    = 'FFF4B084'   # 18:00–21:59  — salmon
COLOR_NIGHT   = 'FFB4C6E7'   # 22:00+       — steel blue
COLOR_OFF     = 'FFD9D9D9'   # grey
COLOR_HOLIDAY = 'FFFFD966'   # yellow
COLOR_SICK    = 'FFFFB3B3'   # light red
COLOR_OTHER   = 'FFFFE4D0'   # light orange
COLOR_WHITE   = 'FFFFFFFF'
COLOR_HEADER  = 'FF1F4E79'   # nav blue for top header
COLOR_HILITE  = 'FF9BC2E6'   # highlights row — same blue as original ODS

THIN   = Side(style='thin',   color='FF000000')
VTHIN  = Side(style='thin',   color='FFD0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SOFT   = Border(left=VTHIN, right=VTHIN, top=VTHIN, bottom=VTHIN)


def _section_bg(section_name):
    """Return ARGB fill colour for a section header."""
    name = (section_name or '').upper()
    for keyword, color in SECTION_HEADER_COLORS.items():
        if keyword in name:
            return color
    return 'FFD6E4F7'   # default light blue


def _section_font_color(bg_argb):
    """Return white text on dark backgrounds, black on light ones."""
    dark = {'FF333333', 'FF00B050', 'FFFF0000', 'FF00B0F0', 'FF808080'}
    return 'FFFFFFFF' if bg_argb in dark else 'FF000000'


def cell_color(shift_val, cell_label=''):
    """Return ARGB background colour for a shift cell."""
    import re
    if cell_label == 'BRUNCH':
        return COLOR_BRUNCH
    if cell_label in ('Breakfast', 'Club'):
        return COLOR_BREAKFAST
    if not shift_val or shift_val in ('OFF', 'OFF/R'):
        return COLOR_OFF
    if shift_val == 'H':
        return COLOR_HOLIDAY
    if shift_val == 'SICK':
        return COLOR_SICK
    if shift_val in ('Comp', 'TBC', 'Paternity', 'Maternity', 'No Work'):
        return COLOR_OTHER
    # Time-band colouring
    m = re.match(r'^(\d{4})-', shift_val)
    if m:
        start = int(m.group(1))
        if start < 800:   return COLOR_EARLY
        if start < 1200:  return COLOR_AM
        if start < 1800:  return COLOR_PM
        if start < 2200:  return COLOR_LATE
        return COLOR_NIGHT
    return COLOR_WHITE


def _write(ws, row, col, value, bg=None, bold=False, size=9,
           halign='center', font_color=None, italic=False, border=True,
           num_format=None):
    """Write a single cell with full formatting. Skips MergedCell slaves."""
    cell = ws.cell(row=row, column=col)
    # openpyxl MergedCell objects are read-only slaves — skip them silently
    if cell.__class__.__name__ == 'MergedCell':
        return cell
    cell.value = value
    fc = font_color or ('FFFFFFFF' if (bg and bg not in (COLOR_WHITE, COLOR_OFF, COLOR_HOLIDAY,
                                                          COLOR_SICK, COLOR_OTHER, COLOR_AM,
                                                          COLOR_EARLY, COLOR_PM)) else 'FF000000')
    # Override for section headers that need white text
    cell.font = Font(name='Calibri', size=size, bold=bold, color=fc, italic=italic)
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
    if border:
        cell.border = BORDER
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if num_format:
        cell.number_format = num_format
    return cell


def export_rota_to_xlsx(rota_period, sections_with_staff, shift_map, dates,
                        events_by_date=None, borrowed_labels=None, shift_labels=None):
    """
    Build an .xlsx rota file.

    Args:
        rota_period:         RotaPeriod ORM object
        sections_with_staff: list of (Section, [Staff, ...]) tuples
        shift_map:           {(staff_id, date_iso): shift_value}
        dates:               list of date objects (up to 14)
        events_by_date:      optional {date: [Event, ...]} for highlights row

    Returns:
        BytesIO containing the .xlsx file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Sanitise sheet name (Excel tab names can't contain: / \ ? * [ ])
    label = str(rota_period.label)
    for ch in r'/\?*[]':
        label = label.replace(ch, '-')
    ws.title = label[:31]

    n_dates = min(len(dates), 14)
    last_data_col = n_dates + 1          # column index of last date col
    total_cols    = last_data_col + 1    # +1 for remarks

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.column_dimensions[get_column_letter(total_cols)].width = 18

    # ── Row 1: notice banner (NO overlapping merges) ──────────────────
    ws.row_dimensions[1].height = 18
    # Just write to individual cells — no merging to avoid MergedCell errors
    _write(ws, 1, 1, '`', bg=COLOR_WHITE, size=8, border=False)
    _write(ws, 1, 2,
           'Please ensure you allow for break times when planning shifts',
           bg='FFFFFF00', italic=True, size=8, halign='left',
           font_color='FF555555', border=False)
    _write(ws, 1, 4,
           'Rota is subjected to change due to business requirement',
           bg=COLOR_WHITE, italic=True, size=8, halign='center',
           font_color='FFCC0000', border=False)

    # ── Row 2: day-name headers ────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    DAY_NAMES = ['Mon', 'Tues', 'Wed', 'Thurs', 'Fri', 'Sat', 'Sun'] * 2
    _write(ws, 2, 1, 'CHEF NAME', bg=COLOR_HEADER, bold=True,
           font_color='FFFFFFFF', size=10, halign='left')
    for i in range(n_dates):
        dn = DAY_NAMES[i] if i < len(DAY_NAMES) else ''
        _write(ws, 2, i + 2, dn, bg=COLOR_HEADER, bold=True,
               font_color='FFFFFFFF', size=9)
    _write(ws, 2, total_cols, 'REMARKS', bg=COLOR_HEADER, bold=True,
           font_color='FFFFFFFF', size=9)

    # ── Row 3: date headers ───────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    _write(ws, 3, 1, rota_period.label, bg=COLOR_HEADER, bold=True,
           font_color='FFFFFFFF', size=9, halign='left')
    for i, d in enumerate(dates[:n_dates]):
        cell = ws.cell(row=3, column=i + 2)
        if cell.__class__.__name__ == 'MergedCell':
            continue
        cell.value = d
        cell.number_format = 'DD/MM/YYYY'
        cell.font = Font(name='Calibri', size=8, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
        cell.border = BORDER

    # ── Rows 4–5: highlights ──────────────────────────────────────────
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 26
    _write(ws, 4, 1, 'UP COMING HIGHLIGHTS', bg=COLOR_HILITE, bold=True,
           size=9, halign='left', font_color='FF1F4E79')
    _write(ws, 5, 1, '', bg=COLOR_HILITE, border=True)

    for i, d in enumerate(dates[:n_dates]):
        evs = (events_by_date or {}).get(d, [])
        # Build a compact highlights string for this date column
        # All events joined — wrap across both rows if needed
        all_txt = '  |  '.join(
            f"{ev.description} ({ev.pax}pax)" if ev.pax else ev.description
            for ev in evs
        )
        # Split roughly in half for rows 4 and 5
        if len(all_txt) > 60 and '  |  ' in all_txt:
            parts = all_txt.split('  |  ')
            mid   = max(1, len(parts) // 2)
            txt4  = '  |  '.join(parts[:mid])
            txt5  = '  |  '.join(parts[mid:])
        else:
            txt4  = all_txt
            txt5  = ''
        _write(ws, 4, i + 2, txt4, bg=COLOR_HILITE, size=8, font_color='FF1F4E79')
        _write(ws, 5, i + 2, txt5, bg=COLOR_HILITE, size=8, font_color='FF1F4E79')

    # ── Row 6: blank spacer ───────────────────────────────────────────
    ws.row_dimensions[6].height = 6

    current_row = 7

    # ── Sections and staff ────────────────────────────────────────────
    for section, staff_members in sections_with_staff:
        sec_bg    = _section_bg(section.name)
        sec_fc    = _section_font_color(sec_bg)

        # Section header: name + day names
        ws.row_dimensions[current_row].height = 16
        _write(ws, current_row, 1, section.name, bg=sec_bg, bold=True,
               size=10, halign='left', font_color=sec_fc)
        for i in range(n_dates):
            _write(ws, current_row, i + 2, DAY_NAMES[i], bg=sec_bg, bold=True,
                   size=8, font_color=sec_fc)
        _write(ws, current_row, total_cols, '', bg=sec_bg)
        current_row += 1

        # Date sub-row
        ws.row_dimensions[current_row].height = 14
        _write(ws, current_row, 1, '', bg=sec_bg)
        for i, d in enumerate(dates[:n_dates]):
            cell = ws.cell(row=current_row, column=i + 2)
            if cell.__class__.__name__ == 'MergedCell':
                continue
            cell.value = d
            cell.number_format = 'DD/MM/YYYY'
            cell.font = Font(name='Calibri', size=8, color=sec_fc)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill('solid', fgColor=sec_bg)
            cell.border = BORDER
        _write(ws, current_row, total_cols, '', bg=sec_bg)
        current_row += 1

        # Staff rows
        for staff in staff_members:
            ws.row_dimensions[current_row].height = 17
            display_name = staff.name
            if getattr(staff, 'role', ''):
                display_name += f' {staff.role}'
            _write(ws, current_row, 1, display_name, bg=COLOR_WHITE,
                   halign='left', size=9, font_color='FF000000')

            for i, d in enumerate(dates[:n_dates]):
                shift    = shift_map.get((staff.id, d.isoformat()), 'OFF')
                cell_lbl = (shift_labels or {}).get((staff.id, d.isoformat()), '')
                bg       = cell_color(shift, cell_lbl)
                borrow    = (borrowed_labels or {}).get((staff.id, d.isoformat()), '')
                # Build cell text: shift on line 1, label on line 2 (if any)
                if cell_lbl and shift not in ('OFF','H','SICK','Comp','Paternity','Maternity','TBC','OFF/R',''):
                    cell_text = f"{shift}\n{cell_lbl}"
                    if borrow:
                        cell_text = f"{shift}\n→{cell_lbl}"
                        bg = 'FFFFF0CC'   # amber tint for borrowed
                else:
                    cell_text = shift
                font_col = 'FFFFFFFF' if cell_lbl == 'BRUNCH' else 'FF000000'
                c = _write(ws, current_row, i + 2, cell_text, bg=bg, size=8,
                           font_color=font_col)
                if cell_lbl and c and c.__class__.__name__ != 'MergedCell':
                    c.alignment = Alignment(horizontal='center', vertical='center',
                                            wrap_text=True)

            _write(ws, current_row, total_cols, '', bg=COLOR_WHITE)
            current_row += 1

        # Blank separator
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    # ── Freeze panes: keep header + first col visible while scrolling ──
    ws.freeze_panes = ws.cell(row=7, column=2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
